import csv
import glob
import os
import zipfile
from os.path import basename

import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook

path_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
files_dir = os.listdir(path_files)
file_archive = os.path.join(path_resources, "archive.zip")


@pytest.fixture()
def create_tmp_folder():
    if os.path.isdir('./tmp') == False:
        os.makedirs('./tmp')


@pytest.fixture()
def clear_dir():
    files = os.path.join(path_resources, '*.*')
    for file in glob.glob(files):
        os.remove(file)


def test_if_have_tmp_folder(create_tmp_folder):
    assert os.path.isdir('./tmp') == True, f"Temp folder does not exist"


def test_create_archive(clear_dir):
    with zipfile.ZipFile(file_archive, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in files_dir:
            add_file = os.path.join(path_files, file)
            zf.write(add_file, basename(add_file))
    files = os.listdir(path_resources)
    assert len(files) == 1, f"Expected number of archive file(s): {len(files)}; actual number of archive file(s): {1}"


def test_csv():
    with zipfile.ZipFile(file_archive) as zf:
        cf = zf.extract("users.csv", './tmp')
        with open(cf) as csvfile:
            csvfile = csv.reader(csvfile)
            list_csv = []
            for r in csvfile:
                text = ",".join(r).replace('\n', " ")
                list_csv.append(text)

            assert list_csv[
                       1] == 'Roland,Brooks,Male,28,r.brooks@randatmail.com,973-1942-40', f"Expected result: {'Roland,Brooks,Male,28,r.brooks@randatmail.com,973-1942-40'}, " \
                                                                                          f"actual result: {list_csv[1]}"
        os.remove(cf)


def test_pdf():
    with zipfile.ZipFile(file_archive) as zf:
        pdf_file = zf.extract('text.pdf', './tmp')
        reader = PdfReader(pdf_file)
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Hello world' in text, f"Expected result: {'Hello world'}; actual result: {text}"
        os.remove(pdf_file)


def test_xlsx():
    with zipfile.ZipFile(file_archive) as zf:
        xf = zf.extract("users.xlsx", './tmp')
        xlsxfile = load_workbook(xf)
        sheet = xlsxfile.active
        assert sheet.cell(row=2, column=1).value == "Roland", f"Expected result: {'Roland'}, " \
                                                              f"actual result: {sheet.cell(row=2, column=1).value}"
        assert sheet.cell(row=2, column=2).value == "Brooks", f"Expected result: {'Brooks'}, " \
                                                              f"actual result: {sheet.cell(row=2, column=2).value}"
        assert sheet.cell(row=2, column=3).value == 'Male', f"Expected result: {'Male'}, " \
                                                            f"actual result: {sheet.cell(row=2, column=3).value}"
        os.remove(xf)
